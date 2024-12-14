import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime

# Функция для применения стиля к ячейкам, включая границы
def apply_cell_style(cell, bold=False, align_center=True, border=None):
    cell.font = Font(bold=bold)
    if align_center:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if border:
        cell.border = border

# Определяем границы для ячеек
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Запрос даты у пользователя
date_str = input('Введите дату в формате дд.мм.гггг: ')

# Проверка корректности введенной даты
try:
    target_date = datetime.strptime(date_str, '%d.%m.%Y')
    print(f"Вы ввели дату: {target_date.strftime('%d.%m.%Y')}")
except ValueError:
    print("Неверный формат даты. Пожалуйста, используйте формат дд.мм.гггг.")
    exit()

inputDir = r'D:\Сменные задания+заявки ОТК'
outputDir = r'D:\Сменные задания+заявки ОТК\Заявки ОТК'

if not os.path.exists(outputDir):
    os.makedirs(outputDir)

output_filename = os.path.join(outputDir, "Заявка_ОТК.xlsx")

# Проверяем, существует ли файл; если нет, создаем новый Workbook
if os.path.exists(output_filename):
    wb = load_workbook(output_filename)
else:
    wb = Workbook()

# Удаляем лист "Sheet", если он пустой и это новый файл
if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
    sheet = wb['Sheet']
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
        del wb['Sheet']

# Создаем новый лист с введённой датой
sheet_name = target_date.strftime('%d.%m')
if sheet_name in wb.sheetnames:
    print(f"Лист с названием {sheet_name} уже существует. Пожалуйста, введите другую дату.")
    exit()

ws = wb.create_sheet(title=sheet_name)

# Устанавливаем заголовок и дату
today_text = f"на «{target_date.strftime('%d')}» _{target_date.strftime('%m')}_ {target_date.strftime('%Y')} года."
ws.merge_cells("A1:F1")
ws["A1"] = "ЗАЯВКА"
ws["A1"].font = Font(bold=True, size=16)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:F2")
ws["A2"] = "на предъявление и сдачу продукции ОТК"
ws["A2"].alignment = Alignment(horizontal="center")

ws.merge_cells("A3:F3")
ws["A3"] = today_text
ws["A3"].alignment = Alignment(horizontal="center")

ws.append([])

# Устанавливаем ширину столбцов
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

# Добавляем заголовки столбцов таблицы
headers = ["Наименование предъявляемой продукции", "№ продукции", "Количество", "Наименование цеха", "Начало работ", "Окончание работ"]
ws.append(headers)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_num, value=header)
    apply_cell_style(cell, bold=True, align_center=True, border=thin_border)

row_index = 6  # Начальная строка для данных

# Функция для парсинга времени
def parse_time(value):
    if isinstance(value, str):
        match = re.match(r'^(\d{1,2}):(\d{2})(?::\d{2})?$', value)
        if match:
            hours, minutes = match.groups()[:2]
            return f"{int(hours):02}:{minutes}"
    return ''

# Чтение данных из исходных папок заводов и заполнение таблицы
factories = ['ЦКТ', 'ЦПМ', 'МСЦ', 'ЭМУ']

data = []  # Для хранения всех строк перед сортировкой

for factory in factories:
    factory_dir = os.path.join(inputDir, factory)

    for root, dirs, files in os.walk(factory_dir):
        for filename in files:
            if filename.startswith('~$'):
                continue
            if filename.endswith(".xlsx"):
                file_path = os.path.join(root, filename)

                try:
                    workbook = load_workbook(file_path, data_only=True)
                    sheet_name = target_date.strftime('%d.%m.%Y')

                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=[2])

                        for _, row in df.iterrows():
                            # Проверяем наличие данных "Сдача ОТК"
                            if pd.notna(row.get('Количество номенклатуры предъявляемая ОТК')):
                                product_name = row['Наименование']
                                product_number = row['№ тепловоза']
                                quantity = row.get('Количество номенклатуры предъявляемая ОТК', 1)
                                workshop = factory

                                # Извлечение времени
                                raw_start_time = str(row.get('План', '')).strip()
                                raw_end_time = str(row.get('Unnamed: 7', '')).strip()

                                start_time = parse_time(raw_start_time)
                                end_time = parse_time(raw_end_time)

                                # Добавляем данные в список для записи
                                data.append([product_name, product_number, quantity, workshop, start_time, end_time])

                except Exception as e:
                    print(f"Ошибка при обработке файла {filename}: {e}")

# Сортировка данных по цеху
data.sort(key=lambda x: x[3])  # Сортируем по столбцу "Наименование цеха"

# Записываем отсортированные данные в Excel
for row in data:
    for col_num, value in enumerate(row, 1):
        cell = ws.cell(row=row_index, column=col_num, value=value)
        apply_cell_style(cell, border=thin_border)
    row_index += 1

# Объединяем ячейки в столбце "Наименование цеха"
def merge_workshop_cells(ws, start_row, end_row, col_index=4):
    current_value = None
    merge_start = start_row

    for row in range(start_row, end_row + 1):
        cell_value = ws.cell(row=row, column=col_index).value

        if cell_value != current_value:
            if current_value is not None and merge_start < row - 1:
                ws.merge_cells(start_row=merge_start, start_column=col_index, end_row=row - 1, end_column=col_index)
                apply_cell_style(ws.cell(row=merge_start, column=col_index), align_center=True, border=thin_border)

            current_value = cell_value
            merge_start = row

    if current_value is not None and merge_start < end_row:
        ws.merge_cells(start_row=merge_start, start_column=col_index, end_row=end_row, end_column=col_index)
        apply_cell_style(ws.cell(row=merge_start, column=col_index), align_center=True, border=thin_border)

merge_workshop_cells(ws, start_row=6, end_row=row_index - 1)

# Добавляем подпись начальника смены в конце
ws.merge_cells(f"A{row_index}:F{row_index}")
ws[f"A{row_index}"] = "Начальник смены"
ws[f"A{row_index}"].alignment = Alignment(horizontal="right")

# Сортируем листы по дате
def sort_worksheets_by_date(wb):
    date_sheets = {}
    for sheet in wb.sheetnames:
        try:
            date_obj = datetime.strptime(sheet, '%d.%m')
            date_sheets[sheet] = date_obj
        except ValueError:
            pass

    sorted_sheets = sorted(date_sheets.items(), key=lambda x: x[1])
    for i, (sheetname, _) in enumerate(sorted_sheets):
        sheet = wb[sheetname]
        wb.move_sheet(sheet, i + 1)

sort_worksheets_by_date(wb)

# Сохранение файла
try:
    wb.save(output_filename)
    print(f"Файл успешно сохранен по пути: {output_filename}")
except Exception as e:
    print(f"Ошибка при сохранении файла: {e}")
