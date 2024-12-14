import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime

# Функция для применения стиля к ячейкам, включая границы
def apply_cell_style_with_borders(cell, bold=False, align_center=True, fill_color=None, border=None):
    cell.font = Font(bold=bold)
    if align_center:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if border:
        cell.border = border

# Функция для форматирования листа Excel
def format_output_sheet(sheet, date_obj, border):
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15

    sheet.merge_cells('A1:G1')
    sheet['A1'] = f"План-задание на {date_obj.strftime('%d.%m.%Y')}"
    apply_cell_style_with_borders(sheet['A1'], bold=True, align_center=True)
    sheet['A1'].font = Font(size=14)

    headers = [
        '№ тепловоза', 'Цех', 'Наименование работы',
        'Процент выполнения работы', 'Начало работ', 'Окончание работ'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num, value=header)
        apply_cell_style_with_borders(cell, bold=True, border=border)

# Функция для проверки значимых строк
def is_significant_row(row):
    return pd.notna(row['№ тепловоза']) and pd.notna(row['Наименование']) and str(row['Наименование']).strip() != ""

# Функция для очистки текста от лишних символов
def clean_text(text):
    if pd.isna(text):
        return ""
    return re.sub(r'[_-]+', '', str(text)).strip()

# Функция для проверки на пустую или нежелательную строку
def is_row_empty_or_unwanted(ws, row_index, num_columns):
    for col in range(1, num_columns + 1):
        value = ws.cell(row=row_index, column=col).value
        if value not in [None, "", "_", "-", "ЦКТ", "ЦПМ", "МСЦ", "ЭМУ"]:
            return False
    return True

# Функция для получения даты с листа Excel
def get_sheet_date(sheet):
    date_str = sheet['A1'].value
    if date_str:
        date_match = re.search(r'\d{2}\.\d{2}\.\d{4}', date_str)
        if date_match:
            try:
                return datetime.strptime(date_match.group(), '%d.%m.%Y')
            except ValueError:
                return None
    return None

# Функция для сортировки листов по дате
def sort_worksheets_by_date(wb):
    date_sheets = {}
    for sheet in wb.sheetnames:
        try:
            date_obj = datetime.strptime(sheet, '%d.%m')
            date_sheets[sheet] = date_obj
        except ValueError:
            pass

    sorted_sheets = sorted(date_sheets.items(), key=lambda x: x[1])
    for i, (sheetname, _) in enumerate(sorted_sheets):
        sheet = wb[sheetname]
        wb.move_sheet(sheet, i + 1)

# Запрос даты у пользователя
date_str = input('Введите нужную дату в формате дд.мм.гггг: ')

try:
    target_date = datetime.strptime(date_str, '%d.%m.%Y')
    print(f"Вы ввели дату: {target_date.strftime('%d.%m.%Y')}")
except ValueError:
    print("Неверный формат даты. Пожалуйста, используйте формат дд.мм.гггг.")
    exit()

inputDir = r'D:\Сменные задания+заявки ОТК'
outputDir = r'D:\Сменные задания+заявки ОТК\План-задание по цехам'

if not os.path.exists(outputDir):
    os.makedirs(outputDir)

thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

factories = ['ЦКТ', 'ЦПМ', 'МСЦ', 'ЭМУ']
plan_filename = os.path.join(outputDir, "План-задание.xlsx")

if os.path.exists(plan_filename):
    wb = load_workbook(plan_filename)
else:
    wb = Workbook()

if 'Sheet' in wb.sheetnames:
    sheet = wb['Sheet']
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
        wb.remove(sheet)

sheet_name = target_date.strftime('%d.%m')
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    wb.remove(ws)

ws = wb.create_sheet(title=sheet_name)

format_output_sheet(ws, target_date, border)

row_index = 3
loco_data = {}

def parse_time(value):
    if isinstance(value, str):
        # Убираем лишние пробелы и проверяем формат "ЧЧ:ММ:СС" или "ЧЧ:ММ"
        match = re.match(r'^(\d{1,2}):(\d{2})(?::\d{2})?$', value)
        if match:
            hours, minutes = match.groups()[:2]
            return f"{int(hours):02}:{minutes}"
    return ''


def is_otk_work(row):
    # Проверяем, заполнен ли столбец "Сдача ОТК"
    otk_value = row.get('Количество номенклатуры предъявляемая ОТК', None)
    return pd.notna(otk_value) and str(otk_value).strip() != ''


# Основной цикл обработки данных
for factory in factories:
    factory_dir = os.path.join(inputDir, factory)
    for root, dirs, files in os.walk(factory_dir):
        for filename in files:
            if filename.startswith('~$'):
                continue

            if filename.endswith(".xlsx"):
                file_path = os.path.join(root, filename)

                try:
                    workbook = load_workbook(file_path, data_only=True)

                    sheet_name = target_date.strftime('%d.%m.%Y')

                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]

                        # Загружаем данные листа в DataFrame
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=[2])

                        for i, r in df.iterrows():
                            if is_significant_row(r):
                                loco_number = clean_text(r['№ тепловоза'])
                                work_name = clean_text(r['Наименование'])

                                # Пропускаем строки, где "Сдача ОТК" не пустая
                                if is_otk_work(r):
                                    continue

                                completion = r.get('Процент выполнения работы', '')
                                if pd.notna(completion) and isinstance(completion, float):
                                    completion = f"{int(completion * 100)}%"
                                else:
                                    completion = ''

                                raw_start_time = str(r.get('План', '')).strip()
                                raw_end_time = str(r.get('Unnamed: 7', '')).strip()

                                start_time = parse_time(raw_start_time)
                                end_time = parse_time(raw_end_time)

                                if loco_number not in loco_data:
                                    loco_data[loco_number] = {}
                                if factory not in loco_data[loco_number]:
                                    loco_data[loco_number][factory] = []
                                loco_data[loco_number][factory].append({
                                    'work_name': work_name,
                                    'completion': completion,
                                    'start_time': start_time,
                                    'end_time': end_time
                                })

                except Exception as e:
                    print(f"Ошибка при обработке файла {filename}: {e}")

# Запись данных в итоговый файл
for loco_number, factory_data in loco_data.items():
    first_row = row_index

    for factory, works in factory_data.items():
        ws.cell(row=row_index, column=2, value=factory)
        apply_cell_style_with_borders(ws.cell(row=row_index, column=2), border=border)
        factory_first_row = row_index

        for work in works:
            ws.cell(row=row_index, column=3, value=work['work_name'])
            apply_cell_style_with_borders(ws.cell(row=row_index, column=3), border=border)
            ws.cell(row=row_index, column=4, value=work['completion'])
            ws.cell(row=row_index, column=5, value=work['start_time'])  # Начало работ
            ws.cell(row=row_index, column=6, value=work['end_time'])  # Окончание работ

            apply_cell_style_with_borders(ws.cell(row=row_index, column=4), border=border)
            apply_cell_style_with_borders(ws.cell(row=row_index, column=5), border=border)
            apply_cell_style_with_borders(ws.cell(row=row_index, column=6), border=border)
            row_index += 1

        if factory_first_row < row_index - 1:
            ws.merge_cells(start_row=factory_first_row, start_column=2, end_row=row_index - 1, end_column=2)
            apply_cell_style_with_borders(ws.cell(row=factory_first_row, column=2), border=border)

    ws.merge_cells(start_row=first_row, start_column=1, end_row=row_index - 1, end_column=1)
    ws.cell(row=first_row, column=1, value=loco_number)
    apply_cell_style_with_borders(ws.cell(row=first_row, column=1), border=border)

for i in range(row_index, 0, -1):
    if is_row_empty_or_unwanted(ws, i, 8):
        ws.delete_rows(i)

sort_worksheets_by_date(wb)

# Сохраняем и выводим подтверждение успешного сохранения
try:
    wb.save(plan_filename)
    print(f"Файл успешно сохранён: {plan_filename}")
except Exception as e:
    print(f"Ошибка при сохранении файла: {e}")